import openpyxl
import json

# Extract data from Excel files
def extract_participants():
    blind_coding = []
    promptverse = []
    
    # Read Blind Coding
    wb1 = openpyxl.load_workbook('BlindCoding Participants.xlsx')
    ws1 = wb1.active
    bc_id = 1
    for row in ws1.iter_rows(min_row=2, values_only=True):
        if row and row[0] and row[1]:
            name = str(row[1]).strip() if row[1] and isinstance(row[1], str) else ""
            if name and not name.startswith("="):
                initials = "".join([word[0].upper() for word in name.split()[:2]])
                blind_coding.append({
                    "id": bc_id,
                    "name": name,
                    "initials": initials or name[0].upper(),
                    "score": 0
                })
                bc_id += 1
    wb1.close()
    
    # Read Promptverse
    wb2 = openpyxl.load_workbook('PromptVerse Participants.xlsx')
    ws2 = wb2.active
    pv_id = 1
    for row in ws2.iter_rows(min_row=2, values_only=True):
        if row and row[0] and row[1]:
            name = str(row[1]).strip() if row[1] and isinstance(row[1], str) else ""
            if name and not name.startswith("="):
                initials = "".join([word[0].upper() for word in name.split()[:2]])
                promptverse.append({
                    "id": pv_id,
                    "name": name,
                    "initials": initials or name[0].upper(),
                    "score": 0
                })
                pv_id += 1
    wb2.close()
    
    return blind_coding, promptverse

# Generate JavaScript code to inject into browser
def generate_setup_code():
    blind_coding, promptverse = extract_participants()
    
    code = f"""
// Paste this in your browser console to load all participants

const blindCodingData = {json.dumps(blind_coding)};
const promptverseData = {json.dumps(promptverse)};

// Save to localStorage
localStorage.setItem('arena-blind-coding-participants', JSON.stringify(blindCodingData));
localStorage.setItem('arena-promptverse-participants', JSON.stringify(promptverseData));
localStorage.setItem('arena-last-participant-id', '{max(len(blind_coding), len(promptverse))}');

// Also save to main state
const state = {{
  blindcoding: blindCodingData,
  promptverse: promptverseData
}};
localStorage.setItem('arena-state', JSON.stringify(state));

console.log('✓ Blind Coding Participants:', blindCodingData.length);
console.log('✓ Promptverse Participants:', promptverseData.length);
console.log('Data loaded successfully!');
location.reload();
"""
    return code

if __name__ == "__main__":
    code = generate_setup_code()
    print(code)
    
    # Also save to a file for easy copying
    with open('setup-data.js', 'w') as f:
        f.write(code)
    print("\n✓ Setup code saved to setup-data.js")
