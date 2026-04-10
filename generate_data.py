import openpyxl

# Extract Blind Coding
blind_coding = []
wb = openpyxl.load_workbook('BlindCoding Participants.xlsx')
ws = wb.active
for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
    if row and row[0] and row[1]:
        name = row[1] if row[1] and isinstance(row[1], str) else (row[0] if isinstance(row[0], str) else '')
        if name and not name.startswith('='):
            name = name.strip()
            initials = ''.join([w[0].upper() for w in name.split()[:2]])
            blind_coding.append({
                "id": idx,
                "name": name,
                "initials": initials[:2],
                "score": 0
            })

# Extract Prompt Verse
promptverse = []
wb2 = openpyxl.load_workbook('PromptVerse Participants.xlsx')
ws2 = wb2.active
for idx, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), 1):
    if row and row[0] and row[1]:
        name = row[1] if row[1] and isinstance(row[1], str) else (row[0] if isinstance(row[0], str) else '')
        if name and not name.startswith('='):
            name = name.strip()
            initials = ''.join([w[0].upper() for w in name.split()[:2]])
            promptverse.append({
                "id": idx,
                "name": name,
                "initials": initials[:2],
                "score": 0
            })

print("const DEFAULT_DATA = {")
print("  blindcoding: [")
for p in blind_coding:
    print(f"    {{ id: {p['id']}, name: '{p['name']}', initials: '{p['initials']}', score: 0 }},")
print("  ],")
print("  promptverse: [")
for p in promptverse:
    print(f"    {{ id: {p['id']}, name: '{p['name']}', initials: '{p['initials']}', score: 0 }},")
print("  ],")
print("}")
